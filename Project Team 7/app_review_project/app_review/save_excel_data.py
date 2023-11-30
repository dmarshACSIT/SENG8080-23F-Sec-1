import openpyxl
from django.core.exceptions import ValidationError
from review_app.models import allAppData

def save_data_from_xlsx():
    # Load the workbook
    final_data_list = []
    workbook = openpyxl.load_workbook("Apps_dataset.xlsx")

    # Select the first worksheet
    print("workbook : {}".format(workbook))
    
    worksheet = workbook.worksheets[0]
    print("WORKSHEET : {}".format(worksheet))

    # Loop through the rows and save each row to the database
    #for row in worksheet.iter_rows(min_row=2):
    for row in worksheet.iter_rows():

        data = {
            'web_scraper_order': row[0].value,
            'web_scraper_order_start_url': row[1].value,
            'app_link_txt': row[2].value,
            'app_link_href': row[3].value,
            'app_name': row[4].value,
            'app_category': row[5].value,
            'app_rating': row[6].value,
            'app_rating_count': row[7].value,
            'number_of_downloads': row[8].value,
            'developer_email': row[9].value,
            'last_update': row[10].value,
            'privacy_policy': row[11].value,
            'contains_ads': row[12].value,
            'in_app_purchase': row[13].value,
            'rated_for': row[14].value,
            'reviews': row[15].value,

            # Add more fields as needed
        }
        final_data_list.append(data)
 
        #Create a new object of MyModel with the data
        obj = allAppData(**data)

        try:
            # Save the object to the database
            obj.full_clean()
            obj.save()
        except ValidationError as e:
            # Handle validation errors here
            print("INSIDE VALUE ERROR!!!!!")
            pass
